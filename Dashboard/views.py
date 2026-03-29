from django.shortcuts import render,HttpResponse, redirect, get_object_or_404
from django.db import models
from .models import *
import uuid
import os
import tempfile
import threading
import re
from openpyxl import Workbook
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import authenticate, login as auth_login, logout
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.hashers import make_password
from django.core.files.images import ImageFile
from math import ceil
from django.core import serializers
from openpyxl import load_workbook
from decimal import Decimal
from time import time
from qrcode import make
from barcode import Code39
from .filters import *
import json
from barcode.writer import ImageWriter
from django.conf import settings
from django.db.models import Q
import io
from django.views.decorators.cache import cache_page
from django.core.cache import cache
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone

def _product_import_cache_key(import_id):
    return f'product_import:{import_id}'

def _get_product_import_state(import_id):
    return cache.get(_product_import_cache_key(import_id))

def _set_product_import_state(import_id, state, timeout_seconds=3600):
    cache.set(_product_import_cache_key(import_id), state, timeout=timeout_seconds)

def _update_product_import_state(import_id, **updates):
    state = _get_product_import_state(import_id) or {}
    state.update(updates)
    _set_product_import_state(import_id, state)
    return state

def _parse_price_to_decimal(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() == 'null':
        return None
    s = s.replace(',', '')
    if '/' in s:
        return None
    if re.fullmatch(r'[+-]?(\d+(\.\d+)?|\.\d+)', s) is None:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None

def _process_product_import_job(import_id, file_path, max_import_rows):
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active

        header_values = [str(c.value).strip().lower() if c.value is not None else '' for c in sheet[1]]
        def find_idx(candidates):
            for name in candidates:
                if name in header_values:
                    return header_values.index(name)
            return None

        idx_parent = find_idx(['parent code','parent_code'])
        idx_child = find_idx(['child code','child_code'])
        idx_location = find_idx(['location'])
        idx_stock = find_idx(['qty','stock','quantity'])
        idx_kpo = find_idx(['kpo'])
        idx_pairing = find_idx(['pairing set','pairing_set','pairing sets'])
        idx_weight = find_idx(['weight'])
        idx_thai_baht = find_idx(['thai baht','thai_baht','thb'])
        idx_usd_rate = find_idx(['usd rate','usd_rate','usd dollar','usd'])
        idx_euro_rate = find_idx(['euro rate','euro_rate','eur'])
        idx_note1 = find_idx(['note 1','note_1'])
        idx_note2 = find_idx(['note 2','note_2'])
        idx_category = find_idx(['category','tag'])
        idx_unit = find_idx(['unit'])
        idx_description = find_idx(['product description','description'])
        idx_images_names = find_idx(['images','image names','images names','image_names'])

        required_indices = [idx_parent, idx_child, idx_location, idx_stock, idx_weight, idx_thai_baht, idx_usd_rate, idx_euro_rate, idx_note1, idx_note2]
        if any(i is None for i in required_indices):
            _update_product_import_state(import_id, status='error', message='Missing required columns in Excel header')
            return

        total_rows = max(0, (sheet.max_row or 1) - 1)
        total_rows = min(total_rows, max_import_rows)
        _update_product_import_state(
            import_id,
            total_rows=total_rows,
            processed_rows=0,
            success_count=0,
            failed_count=0,
            skipped_count=0,
            failed_rows=[],
            failed_rows_truncated=False,
            failed_rows_truncated_count=0,
            status='running',
            message=''
        )

        processed_rows = 0
        success_count = 0
        failed_count = 0
        skipped_count = 0
        failed_rows = []
        failed_rows_limit = 5000
        failed_rows_truncated_count = 0

        for row_index, fields in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            if row_index > max_import_rows:
                break

            processed_rows = row_index
            _update_product_import_state(import_id, processed_rows=processed_rows)

            try:
                if not fields or (idx_parent is not None and not fields[idx_parent]):
                    skipped_count += 1
                    continue

                parent_code = fields[idx_parent]
                child_code = fields[idx_child]
                location = fields[idx_location]
                stock = fields[idx_stock]
                kpo = fields[idx_kpo] if idx_kpo is not None else None

                if parent_code is None or str(parent_code).strip() == '':
                    raise ValueError('Missing Parent Code')
                if child_code is None or str(child_code).strip() == '':
                    raise ValueError('Missing Child Code')
                if location is None or str(location).strip() == '':
                    raise ValueError('Missing Location')

                pairing_sets_data = None
                if idx_pairing is not None and fields[idx_pairing]:
                    pairing_sets_data = str(fields[idx_pairing]).split(',')

                try:
                    weight = Decimal(str(fields[idx_weight])) if fields[idx_weight] is not None else Decimal('0.00')
                except (ValueError, TypeError):
                    weight = Decimal('0.00')

                thai_baht = fields[idx_thai_baht]
                usd_rate = fields[idx_usd_rate]
                euro_rate = fields[idx_euro_rate]
                note_1 = fields[idx_note1]
                note_2 = fields[idx_note2]
                description = fields[idx_description] if idx_description is not None else None
                unit = fields[idx_unit] if idx_unit is not None else None
                category_name = fields[idx_category] if idx_category is not None else None

                images_names_data = None
                if idx_images_names is not None and len(fields) > idx_images_names and fields[idx_images_names]:
                    images_names_data = str(fields[idx_images_names]).split(',')

                child_code_str = str(child_code).strip()
                product = Product.objects.filter(child_code=child_code_str).first()
                if product:
                    product.location = location
                    product.stock = stock
                    product.kpo = kpo
                    product.weight = weight
                    product.thai_baht = thai_baht
                    product.usd_rate = usd_rate
                    product.euro_rate = euro_rate
                    product.note_1 = note_1
                    product.note_2 = note_2
                    if description is not None:
                        product.description = description
                    if unit is not None:
                        product.unit = unit
                    if category_name:
                        tag_obj, _ = Tag.objects.get_or_create(name=str(category_name).strip())
                        product.tag = tag_obj
                    product.save()
                else:
                    product = Product.objects.create(
                        parent_code=str(parent_code).strip(),
                        child_code=child_code_str,
                        location=str(location).strip(),
                        stock=stock,
                        kpo=kpo,
                        weight=weight,
                        thai_baht=thai_baht,
                        usd_rate=usd_rate,
                        euro_rate=euro_rate,
                        note_1=note_1,
                        note_2=note_2,
                        description=description or None,
                        unit=unit or None
                    )
                    if category_name:
                        try:
                            tag_obj, _ = Tag.objects.get_or_create(name=str(category_name).strip())
                            product.tag = tag_obj
                            product.save()
                        except Exception:
                            pass

                pairing_sets = []
                if pairing_sets_data:
                    for pair_value in pairing_sets_data:
                        pair_value_str = str(pair_value).strip()
                        if not pair_value_str:
                            continue
                        pairing_set_obj, created = PairingSet.objects.get_or_create(pair_value=pair_value_str)
                        pairing_sets.append(pairing_set_obj)
                    product.pairing_set.set(pairing_sets)

                if images_names_data:
                    images_names = []
                    for image_name in images_names_data:
                        image_name_str = str(image_name).strip()
                        if not image_name_str:
                            continue
                        image_name_obj, created = ImageName.objects.get_or_create(name=image_name_str)
                        images_names.append(image_name_obj)
                    product.images_names.set(images_names)
                    image_files = Image.objects.filter(image__in=[f'product_images/{img_name.name}' for img_name in images_names])
                    product.images.set(image_files)
                    product.save()

                if not product.qrcode_image:
                    qr_img = make(child_code_str)
                    qr_img_name = f'{str(time())}.png'
                    buffer = io.BytesIO()
                    qr_img.save(buffer, format='PNG')
                    buffer.seek(0)
                    product.qrcode_image.save(qr_img_name, ImageFile(buffer), save=True)

                if not product.barcode_image:
                    barcode_obj = Code39(child_code_str, writer=ImageWriter())
                    buffer = io.BytesIO()
                    barcode_obj.write(buffer)
                    buffer.seek(0)
                    product.barcode_image.save(f'{child_code_str}.png', ImageFile(buffer), save=True)

                product.save()
                success_count += 1
            except Exception as e:
                failed_count += 1
                excel_row_number = row_index + 1
                err_text = str(e) if str(e) else e.__class__.__name__
                if len(err_text) > 300:
                    err_text = err_text[:300]
                child_code_value = None
                try:
                    if fields and idx_child is not None and len(fields) > idx_child:
                        child_code_value = fields[idx_child]
                except Exception:
                    child_code_value = None

                if len(failed_rows) < failed_rows_limit:
                    failed_rows.append({
                        'row': excel_row_number,
                        'child_code': (str(child_code_value).strip() if child_code_value is not None else ''),
                        'error': err_text
                    })
                else:
                    failed_rows_truncated_count += 1

                _update_product_import_state(
                    import_id,
                    success_count=success_count,
                    failed_count=failed_count,
                    skipped_count=skipped_count,
                    failed_rows=failed_rows,
                    failed_rows_truncated=(failed_rows_truncated_count > 0),
                    failed_rows_truncated_count=failed_rows_truncated_count
                )

            if row_index % 25 == 0:
                _update_product_import_state(
                    import_id,
                    success_count=success_count,
                    failed_count=failed_count,
                    skipped_count=skipped_count,
                    failed_rows=failed_rows,
                    failed_rows_truncated=(failed_rows_truncated_count > 0),
                    failed_rows_truncated_count=failed_rows_truncated_count
                )

        _update_product_import_state(
            import_id,
            status='done',
            processed_rows=processed_rows,
            success_count=success_count,
            failed_count=failed_count,
            skipped_count=skipped_count,
            failed_rows=failed_rows,
            failed_rows_truncated=(failed_rows_truncated_count > 0),
            failed_rows_truncated_count=failed_rows_truncated_count,
            message='Import completed'
        )
    except Exception as e:
        _update_product_import_state(import_id, status='error', message=str(e))
    finally:
        try:
            try:
                wb.close()
            except Exception:
                pass
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            pass

@login_required
def product_import_start(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

    xlsx_file = request.FILES.get('xlsx_file')
    if not xlsx_file:
        return JsonResponse({'success': False, 'message': 'No file uploaded'}, status=400)

    if not (xlsx_file.name.endswith('.xlsx') or xlsx_file.name.endswith('.xls')):
        return JsonResponse({'success': False, 'message': 'File is not xlsx type'}, status=400)

    max_import_rows = 10000
    import_id = uuid.uuid4().hex

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(xlsx_file.name)[1] or '.xlsx')
    tmp_path = tmp.name
    try:
        for chunk in xlsx_file.chunks():
            tmp.write(chunk)
    finally:
        tmp.close()

    try:
        wb = load_workbook(tmp_path, data_only=True, read_only=True)
        sheet = wb.active
        total_rows = max(0, (sheet.max_row or 1) - 1)
        wb.close()
    except Exception:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
        return JsonResponse({'success': False, 'message': 'Error reading Excel file'}, status=400)

    if total_rows > max_import_rows:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
        return JsonResponse({'success': False, 'message': f'Import limit is {max_import_rows} rows per upload. Please split the file and try again.'}, status=400)

    _set_product_import_state(import_id, {
        'import_id': import_id,
        'total_rows': total_rows,
        'processed_rows': 0,
        'success_count': 0,
        'failed_count': 0,
        'skipped_count': 0,
        'failed_rows': [],
        'failed_rows_truncated': False,
        'failed_rows_truncated_count': 0,
        'status': 'queued',
        'message': ''
    })

    t = threading.Thread(target=_process_product_import_job, args=(import_id, tmp_path, max_import_rows), daemon=True)
    t.start()

    _update_product_import_state(import_id, status='running')
    return JsonResponse({'success': True, 'import_id': import_id, 'total_rows': total_rows, 'max_rows': max_import_rows})

@login_required
def product_import_status(request):
    import_id = request.GET.get('import_id')
    if not import_id:
        return JsonResponse({'success': False, 'message': 'Missing import_id'}, status=400)

    state = _get_product_import_state(import_id)
    if not state:
        return JsonResponse({'success': False, 'message': 'Import not found'}, status=404)

    return JsonResponse({'success': True, **state})


@csrf_exempt
@login_required()
def upload_bulk_images(request):
    if request.method == 'POST':
        uploaded_images = request.FILES.getlist("file")
        if not uploaded_images:
            return JsonResponse({'success': False, 'message': 'No images selected!'})
        
        uploaded_count = 0
        linked_count = 0
        results = []
        
        for uploaded_image in uploaded_images:
            try:
                # Check if image already exists
                image, created = Image.objects.get_or_create(
                    defaults={'image': uploaded_image},
                    image__icontains=uploaded_image.name.split('.')[0]
                )
                
                if created:
                    uploaded_count += 1
                    # Try to link with products based on image name
                    image_name_without_ext = uploaded_image.name.split('.')[0]
                    products = Product.objects.filter(
                        Q(parent_code__icontains=image_name_without_ext) |
                        Q(child_code__icontains=image_name_without_ext)
                    )
                    
                    linked_products = []
                    for product in products:
                        if not product.images.filter(id=image.id).exists():
                            product.images.add(image)
                            product.save()
                            
                            # Create persistent link
                            ProductImageLink.objects.get_or_create(
                                image=image,
                                parent_code=product.parent_code,
                                child_code=product.child_code
                            )
                            
                            linked_products.append(f"{product.parent_code}-{product.child_code}")
                            linked_count += 1
                    
                    results.append({
                        'name': uploaded_image.name,
                        'status': 'uploaded',
                        'linked_products': linked_products
                    })
                else:
                    results.append({
                        'name': uploaded_image.name,
                        'status': 'already_exists',
                        'linked_products': []
                    })
            except Exception as e:
                results.append({
                    'name': uploaded_image.name,
                    'status': 'error',
                    'error': str(e),
                    'linked_products': []
                })
        
        return JsonResponse({
            'success': True, 
            'message': f'Successfully uploaded {uploaded_count} images and linked to {linked_count} products!',
            'results': results,
            'uploaded_count': uploaded_count,
            'linked_count': linked_count
        })


@csrf_exempt
@login_required()
def image_management(request):
    """View for managing images and linking them to products"""
    if request.method == 'GET':
        # Get all images with their linked products
        images = Image.objects.prefetch_related('product_set').all().order_by('-id')
        products = Product.objects.all().order_by('child_code')
        
        # Prepare image data with linked products
        image_data = []
        for image in images:
            linked_products = image.product_set.all()
            image_data.append({
                'id': image.id,
                'name': image.image.name.split('/')[-1] if image.image else 'No name',
                'url': image.image.url if image.image else '',
                'linked_products': [{
                    'id': p.id,
                    'parent_code': p.parent_code,
                    'child_code': p.child_code
                } for p in linked_products]
            })
        
        return render(request, 'image_management.html', {
            'images': image_data,
            'products': products
        })
    
    elif request.method == 'POST':
        data = json.loads(request.body)
        action = data.get('action')
        
        if action == 'link_image':
            image_id = data.get('image_id')
            product_ids = data.get('product_ids', [])
            
            try:
                image = Image.objects.get(id=image_id)
                products = Product.objects.filter(id__in=product_ids)
                
                # Add new product links without clearing existing ones
                for product in products:
                    if not product.images.filter(id=image.id).exists():
                        product.images.add(image)
                    
                    # Create persistent link
                    ProductImageLink.objects.get_or_create(
                        image=image,
                        parent_code=product.parent_code,
                        child_code=product.child_code
                    )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Image linked to {len(products)} products successfully!'
                })
            except Image.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Image not found!'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Error linking image: {str(e)}'
                })
        
        elif action == 'unlink_image':
            image_id = data.get('image_id')
            product_id = data.get('product_id')
            
            try:
                image = Image.objects.get(id=image_id)
                product = Product.objects.get(id=product_id)
                product.images.remove(image)
                
                # Remove persistent link only if manually unlinked
                ProductImageLink.objects.filter(
                    image=image,
                    parent_code=product.parent_code,
                    child_code=product.child_code
                ).delete()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Image unlinked successfully!'
                })
            except (Image.DoesNotExist, Product.DoesNotExist):
                return JsonResponse({
                    'success': False,
                    'message': 'Image or product not found!'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Error unlinking image: {str(e)}'
                })
        
        elif action == 'delete_image':
            image_id = data.get('image_id')
            
            try:
                image = Image.objects.get(id=image_id)
                image.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Image deleted successfully!'
                })
            except Image.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Image not found!'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Error deleting image: {str(e)}'
                })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required()
def form_view(request):
    if request.method == 'POST':
        data = request.POST
        type = data.get('type')
        if type == 'bulk-images':
            uploaded_images = request.FILES.getlist("images") 
            if not uploaded_images:
                messages.error(request,"Select at least one image!")
                return redirect('form')    
            for uploaded_image in uploaded_images:
                # Check if image already exists by filename
                existing_image = Image.objects.filter(image__icontains=uploaded_image.name).first()
                if existing_image:
                    image = existing_image
                    created = False
                else:
                    image = Image.objects.create(image=uploaded_image)
                    created = True
                
                products = Product.objects.all()
                for product in products:
                    if product.images_names.filter(name=uploaded_image.name).exists():
                        product.images.add(image)
                        product.save()
            messages.success(request,"Successfully uploaded images!")
            return redirect('form')
        if type == 'bulk-create':
            try:
                xlsx_file = request.FILES.get("xlsx_file")
                if not (xlsx_file.name.endswith('.xlsx') or xlsx_file.name.endswith('.xls')):
                    messages.error(request,'File is not xlsx type')
                    return redirect("form")

                max_import_rows = 10000
                wb = load_workbook(xlsx_file, data_only=True)
                sheet = wb.active
                if sheet.max_row and (sheet.max_row - 1) > max_import_rows:
                    messages.error(request, f'Import limit is {max_import_rows} rows per upload. Please split the file and try again.')
                    return redirect("form")
                header_values = [str(c.value).strip().lower() if c.value is not None else '' for c in sheet[1]]
                def find_idx(candidates):
                    for name in candidates:
                        if name in header_values:
                            return header_values.index(name)
                    return None
                idx_parent = find_idx(['parent code','parent_code'])
                idx_child = find_idx(['child code','child_code'])
                idx_location = find_idx(['location'])
                idx_stock = find_idx(['qty','stock','quantity'])
                idx_kpo = find_idx(['kpo'])
                idx_pairing = find_idx(['pairing set','pairing_set','pairing sets'])
                idx_weight = find_idx(['weight'])
                idx_thai_baht = find_idx(['thai baht','thai_baht','thb'])
                idx_usd_rate = find_idx(['usd rate','usd_rate','usd dollar','usd'])
                idx_euro_rate = find_idx(['euro rate','euro_rate','eur'])
                idx_note1 = find_idx(['note 1','note_1'])
                idx_note2 = find_idx(['note 2','note_2'])
                idx_category = find_idx(['category','tag'])
                idx_unit = find_idx(['unit'])
                idx_description = find_idx(['product description','description'])
                idx_images_names = find_idx(['images','image names','images names','image_names'])
                required_indices = [idx_parent, idx_child, idx_location, idx_stock, idx_weight, idx_thai_baht, idx_usd_rate, idx_euro_rate, idx_note1, idx_note2]
                if any(i is None for i in required_indices):
                    messages.error(request,'Missing required columns in Excel header')
                    return redirect("form")
                for row_index, fields in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                    if row_index > max_import_rows:
                        break
                    if not fields or (idx_parent is not None and not fields[idx_parent]):
                        continue
                    parent_code = fields[idx_parent]
                    child_code = fields[idx_child]
                    location = fields[idx_location]
                    stock = fields[idx_stock]
                    kpo = fields[idx_kpo] if idx_kpo is not None else None
                    pairing_sets_data = None
                    if idx_pairing is not None and fields[idx_pairing]:
                        pairing_sets_data = str(fields[idx_pairing]).split(',')
                    try:
                        weight = Decimal(str(fields[idx_weight])) if fields[idx_weight] is not None else Decimal('0.00')
                    except (ValueError, TypeError):
                        weight = Decimal('0.00')
                    thai_baht = fields[idx_thai_baht]
                    usd_rate = fields[idx_usd_rate]
                    euro_rate = fields[idx_euro_rate]
                    note_1 = fields[idx_note1]
                    note_2 = fields[idx_note2]
                    description = fields[idx_description] if idx_description is not None else None
                    unit = fields[idx_unit] if idx_unit is not None else None
                    category_name = fields[idx_category] if idx_category is not None else None
                    images_names_data = None
                    if idx_images_names is not None and len(fields) > idx_images_names and fields[idx_images_names]:
                        images_names_data = str(fields[idx_images_names]).split(',')
                    product =  Product.objects.filter(child_code=child_code).first()
                    if product:
                        product.location = location
                        product.stock = stock
                        product.kpo = kpo
                        product.weight = weight
                        product.thai_baht = thai_baht
                        product.usd_rate = usd_rate
                        product.euro_rate = euro_rate
                        product.note_1 = note_1
                        product.note_2 = note_2
                        if description is not None:
                            product.description = description
                        if unit is not None:
                            product.unit = unit
                        if category_name:
                            tag_obj, _ = Tag.objects.get_or_create(name=str(category_name).strip())
                            product.tag = tag_obj
                        product.save()
                    else:      
                        product = Product.objects.create(parent_code=parent_code, child_code=child_code, location=location, stock=stock, kpo=kpo, weight=weight, thai_baht=thai_baht, usd_rate=usd_rate, euro_rate=euro_rate, note_1=note_1, note_2=note_2, description=description or None, unit=unit or None)
                        if category_name:
                            try:
                                tag_obj, _ = Tag.objects.get_or_create(name=str(category_name).strip())
                                product.tag = tag_obj
                                product.save()
                            except Exception:
                                pass
                    pairing_sets = []
                    if pairing_sets_data:
                        for pair_value in pairing_sets_data:
                            pairing_set_obj, created = PairingSet.objects.get_or_create(pair_value=str(pair_value).strip())
                            pairing_sets.append(pairing_set_obj)
                        product.pairing_set.set(pairing_sets)
                    if images_names_data:
                        images_names = []
                        for image_name in images_names_data:
                            image_name_obj, created = ImageName.objects.get_or_create(name=str(image_name).strip())
                            images_names.append(image_name_obj)
                        product.images_names.set(images_names)
                        image_files = Image.objects.filter(image__in=[f'product_images/{img_name.name}' for img_name in images_names])
                        product.images.set(image_files)
                        product.save()

                    if not product.qrcode_image:
                        qr_img = make(child_code)
                        qr_img_name = f'{str(time())}.png'
                        buffer = io.BytesIO()
                        qr_img.save(buffer, format='PNG')
                        buffer.seek(0)
                        product.qrcode_image.save(qr_img_name, ImageFile(buffer), save=True)

                    if not product.barcode_image:
                        barcode_obj = Code39(child_code, writer=ImageWriter())
                        buffer = io.BytesIO()
                        barcode_obj.write(buffer)
                        buffer.seek(0)
                        product.barcode_image.save(f'{child_code}.png', ImageFile(buffer), save=True)
                    
                    product.save()
                    
            except Exception as e:
                messages.error(request, f"Error processing Excel file: {str(e)}. Please check your file format and data.")
                return redirect('form')
                
            messages.success(request,"Successfully saved data from file!")
            return redirect('form')
        
        if type == 'product-grouping':
            tag = data.get('tag')
            products_ids = request.POST.getlist('products')
            tag_obj, created = Tag.objects.get_or_create(name=tag)
            products = Product.objects.filter(id__in=products_ids).update(tag=tag_obj)
            messages.success(request,"Successfully Grouping!")
            return redirect('form')

        if type == 'create':
            parent_code = data.get('parent_code')
            child_code = data.get('child_code')
            location = data.get('location')
            stock = data.get('stock')
            kpo = data.get('kpo')
            status = data.get('status')
            pairing_set_ids = request.POST.getlist('pairing_set')
            weight = data.get('weight')
            thai_baht = data.get('thai_baht')
            usd_rate = data.get('usd_rate')
            euro_rate = data.get('euro_rate')
            note_1 = data.get('note_1')
            note_2 = data.get('note_2')
            description = data.get('description')
            unit = data.get('unit')
            tag_id = data.get('tag_id')
            uploaded_images = request.FILES.getlist("images") 

            product_check =  Product.objects.filter(child_code=child_code).first()
            if product_check:
                messages.error(request, "Product with this child code already exists!")
                return redirect('form')

            product = Product.objects.create(parent_code=parent_code, child_code=child_code, location=location, stock=stock, kpo=kpo, weight=weight, thai_baht=thai_baht, usd_rate=usd_rate, euro_rate=euro_rate, note_1=note_1, note_2=note_2, description=description, unit=unit)
            product.pairing_set.add(*pairing_set_ids)
            if tag_id:
                try:
                    product.tag = Tag.objects.get(id=int(tag_id))
                    product.save()
                except Exception:
                    pass

            if uploaded_images:
                for uploaded_image in uploaded_images:
                    image, created = Image.objects.get_or_create(image__exact=uploaded_image.name)
                    if created:
                        image.image = uploaded_image
                        image.save()
                    product.images.add(image)
                    product.save()

            qr_img = make(child_code)
            qr_img_name = f'{str(time())}.png'
            buffer = io.BytesIO()
            qr_img.save(buffer, format='PNG')
            buffer.seek(0)
            product.qrcode_image.save(qr_img_name, ImageFile(buffer), save=True)

            barcode_obj = Code39(child_code, writer=ImageWriter())
            buffer = io.BytesIO()
            barcode_obj.write(buffer)
            buffer.seek(0)
            product.barcode_image.save(f'{child_code}.png', ImageFile(buffer), save=True)

            product.save()
            messages.success(request,"Successfully Created!")
            return redirect('form')
        if type == 'update':
            pro_id = data.get('pro_id')
            parent_code = data.get('parent_code')
            child_code = data.get('child_code')
            location = data.get('location')
            stock = data.get('stock')
            kpo = data.get('kpo')
            status = data.get('status')
            pairing_set_ids = request.POST.getlist('pairing_set')
            weight = data.get('weight')
            thai_baht = data.get('thai_baht')
            usd_rate = data.get('usd_rate')
            euro_rate = data.get('euro_rate')
            note_1 = data.get('note_1')
            note_2 = data.get('note_2')
            description = data.get('description')
            unit = data.get('unit')
            tag_id = data.get('tag_id')

            product_check =  Product.objects.filter(child_code=child_code).exclude(id=pro_id).first()
            if product_check:
                messages.error(request, "Product with this child code already exists!")
                return redirect('form')
            
            product = Product.objects.get(id=pro_id)
            
            product.parent_code=parent_code
            product.child_code=child_code 
            product.location=location 
            product.stock=stock 
            product.kpo=kpo 
            product.weight=weight 
            product.thai_baht=thai_baht 
            product.usd_rate=usd_rate 
            product.euro_rate=euro_rate 
            product.note_1=note_1
            product.note_2=note_2
            product.description=description
            product.unit=unit
            product.pairing_set.clear()
            product.pairing_set.add(*pairing_set_ids)
            if tag_id:
                try:
                    product.tag = Tag.objects.get(id=int(tag_id))
                except Exception:
                    product.tag = None
            
            # Handle image uploads for updates
            uploaded_images = request.FILES.getlist("images")
            if uploaded_images:
                for uploaded_image in uploaded_images:
                    image, created = Image.objects.get_or_create(image__exact=uploaded_image.name)
                    if created:
                        image.image = uploaded_image
                        image.save()
                    product.images.add(image)
                    product.save()

            qr_img = make(child_code)
            qr_img_name = f'{str(time())}.png'
            buffer = io.BytesIO()
            qr_img.save(buffer, format='PNG')
            buffer.seek(0)
            product.qrcode_image.save(qr_img_name, ImageFile(buffer), save=True)

            barcode_obj = Code39(child_code, writer=ImageWriter())
            buffer = io.BytesIO()
            barcode_obj.write(buffer)
            buffer.seek(0)
            product.barcode_image.save(f'{child_code}.png', ImageFile(buffer), save=True)
            
            product.save()
            messages.success(request,"Successfully Updated!")
            return redirect('form')
        elif type == 'delete':
            pro_id = data.get('product_id')
            Product.objects.get(id=pro_id).delete()
            messages.success(request,"Successfully Deleted!")
            return redirect('form')
        elif type == 'delete_all_products':
            Product.objects.all().delete()
            messages.success(request,"Successfully deleted all products!")
            return redirect('form')
    elif request.method == 'DELETE':
        data = json.loads(request.body)
        if data.get('type') == 'bulk-delete':
            product_ids = data.get('ids', [])
            if product_ids:
                Product.objects.filter(id__in=product_ids).delete()
                return JsonResponse({'success': True, 'message': 'Selected products deleted successfully'})
            return JsonResponse({'success': False, 'message': 'No products selected'})
    elif request.method == "GET":
        get_req_type = request.GET.get('get_req_type','all')
        if get_req_type == "individual":
            pro_id = request.GET.get('pro_id')
            pro_obj = Product.objects.prefetch_related('pairing_set', 'images').filter(id=pro_id).first()
            pro_json = {
                'parent_code':pro_obj.parent_code,
                'child_code':pro_obj.child_code,
                'location':pro_obj.location,
                'stock':pro_obj.stock,
                'kpo':pro_obj.kpo,
                'weight':pro_obj.weight,
                'thai_baht':pro_obj.thai_baht,
                'usd_rate':pro_obj.usd_rate,
                'euro_rate':pro_obj.euro_rate,
                'note_1': pro_obj.note_1,
                'note_2': pro_obj.note_2,
                'description': pro_obj.description,
                'unit': pro_obj.unit,
                'tag_id': pro_obj.tag.id if pro_obj.tag else None,
                'tag_name': pro_obj.tag.name if pro_obj.tag else '',
                'pairing_set':[ps.id for ps in pro_obj.pairing_set.all()],
                'images': [{'url': image.image.url} for image in pro_obj.images.all()]
            }
            return JsonResponse(pro_json,status=200)
        elif get_req_type == "all":
            pairing_set_obj = PairingSet.objects.all().order_by('-id')
            tags_obj = Tag.objects.all().order_by('name')
            
        return render(request, 'form.html', {
            'pairing_set' : pairing_set_obj,
            'tags': tags_obj,
        })

@login_required
def product_api(request):
    # Handle sorting
    sort_by = request.GET.get('sort_by', 'id')
    sort_order = request.GET.get('sort_order', 'desc')
    
    # Map frontend sort fields to model fields
    sort_field_mapping = {
        'parent_code': 'parent_code',
        'child_code': 'child_code',
        'location': 'location',
        'kpo': 'kpo',
        'price': 'thai_baht',
        'stock': 'stock',
        'weight': 'weight',
        'id': 'id'
    }
    
    # Get the actual field name
    actual_sort_field = sort_field_mapping.get(sort_by, 'id')
    
    # Apply sort order
    if sort_order == 'desc':
        order_by = f'-{actual_sort_field}'
    else:
        order_by = actual_sort_field
    
    product_obj = Product.objects.prefetch_related('images_names', 'images', 'pairing_set').order_by(order_by)
    product_obj = ProductFilter(request.GET, queryset=product_obj)
    data = [{
        'id': pro_obj.id,
        'image':pro_obj.images.first().image.url if pro_obj.images.exists() else '',
        'parent_code':pro_obj.parent_code,
        'child_code':pro_obj.child_code,
        'location':pro_obj.location,
        'stock':pro_obj.stock,
        'kpo':pro_obj.kpo,
        'weight':pro_obj.weight,
        'thai_baht':pro_obj.thai_baht,
        'usd_rate':pro_obj.usd_rate,
        'euro_rate':pro_obj.euro_rate,
        'note_1': pro_obj.note_1,
        'note_2': pro_obj.note_2,
        'description': pro_obj.description,
        'unit': pro_obj.unit,
        'tag':pro_obj.tag.name if pro_obj.tag else '',
        'pairing_set':[ps.id for ps in pro_obj.pairing_set.all()],
        'image_count': pro_obj.images.count()
    } for pro_obj in product_obj.qs]

    data = custom_paginate_queryset(
        queryset=data,
        page=int(request.GET.get('page', 1)),
        page_size=int(request.GET.get('page_size', 10))
    )
    return JsonResponse(data, safe=False)



@login_required()     
def product_detail(request, id):
    prod = Product.objects.get(id=id)
    return render(request, 'product_detail.html', {'prod' : prod})

# @cache_page(60 * 60 * 24)  # Cache for one day (24 hours)
@csrf_exempt
def single_product(request):
    if request.method == 'GET':
        # Support multiple search parameters
        child_code = request.GET.get('child_code')
        parent_code = request.GET.get('parent_code')
        code = request.GET.get('code')  # Generic code parameter for QR/barcode scanning
        
        # If generic 'code' parameter is provided, search in both child_code and parent_code
        if code:
            prod = Product.objects.prefetch_related('pairing_set', 'images').filter(
                Q(child_code=code) | Q(parent_code=code)
            ).first()
        elif child_code:
            prod = Product.objects.prefetch_related('pairing_set', 'images').filter(child_code=child_code).first()
        elif parent_code:
            prod = Product.objects.prefetch_related('pairing_set', 'images').filter(parent_code=parent_code).first()
        else:
            return JsonResponse({'error': 'Please provide code, child_code, or parent_code parameter'}, status=400)
            
        if prod:
            # Get matching items (pairing set products)
            matching_items = []
            for pairing in prod.pairing_set.all():
                matching_products = Product.objects.prefetch_related('images').filter(
                    pairing_set__pair_value=pairing.pair_value
                ).exclude(id=prod.id)[:5]  # Limit to 5 matching items
                
                for matching_prod in matching_products:
                    matching_items.append({
                        'code': matching_prod.child_code,
                        'parent_code': matching_prod.parent_code,
                        'image_url': matching_prod.images.first().image.url if matching_prod.images.exists() else None
                    })
            
            prod_data = {
                'success': True,
                'id': prod.id,
                'parent_code': prod.parent_code,
                'child_code': prod.child_code,
                'location': prod.location,
                'stock': prod.stock,
                'kpo': prod.kpo,
                'pairing_set': [{'value': pairing.pair_value} for pairing in prod.pairing_set.all()],
                'weight': str(prod.weight),
                'thai_baht': prod.thai_baht,
                'usd_rate': prod.usd_rate,
                'euro_rate': prod.euro_rate,
                'note_1': prod.note_1,
                'note_2': prod.note_2,
                'description': prod.description,
                'unit': prod.unit,
                'tag': prod.tag.name if prod.tag else '',
                'images': [{'url': request.build_absolute_uri(image.image.url)} for image in prod.images.all()],
                'qrcode_image': request.build_absolute_uri(prod.qrcode_image.url) if prod.qrcode_image else None,
                'barcode_image': request.build_absolute_uri(prod.barcode_image.url) if prod.barcode_image else None,
                'matching_items': matching_items,
                'created_at': prod.created_at.isoformat() if prod.created_at else None,
                'updated_at': prod.updated_at.isoformat() if prod.updated_at else None,
            }
            return JsonResponse(prod_data, safe=False)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Product not found',
                'error': 'No product found with the provided code'
            }, status=404)
    return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)
    
def login_view(request):
    if request.user.is_authenticated:
            return redirect("form")
    if request.method == 'POST':
        data = request.POST
        username = data['username']
        password = data['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)
            messages.success(request,"Successfully Login!")
            return redirect("form")
        else:
            messages.warning(request,"Incorrect email or password!")        
    return render(request, 'login.html')

@login_required()
def change_password(request):
    if request.method == "POST":
        p1 = request.POST.get('password1')
        p2 = request.POST.get('password2')
        user = User.objects.get(username=request.user.username)
        if p1 == p2:
            user.set_password(p1) 
            user.save()   
            messages.success(request,"Successfully Change Password!")
            return redirect('form')
        else:
            messages.error(request,"Password does not match!")
            return redirect('change_password')
    return render(request, 'change_password.html')

@login_required()
def user_logout(request):
    try:
        logout(request)
        messages.success(request, 'Logout Successfully!')
        return redirect('form')
    except:
            messages.warning(request, 'Request is not responed please check your internet connection and try again!')
            return redirect('form')

def custom_paginate_queryset(queryset, page, page_size):
    total_items = len(queryset)
    total_pages = ceil(total_items / page_size)
    start_index = (page - 1) * page_size
    end_index = start_index + page_size
    paginated_queryset = queryset[start_index:end_index]
    
    return {
        'pagination': {
            'next': page + 1 if page < total_pages else None,
            'previous': page - 1 if page > 1 else None,
            'count': total_items,
            'total_pages': total_pages,
            'current_page': page,
        },
        'results': paginated_queryset
    }


def export_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Karen Data'

    columns = ['Parent Code', 'Child Code', 'Location', 'QTY', 'kpo', 'pairing_set', 'weight', 'thai_baht', 'usd_rate', 'euro_rate', 'Category', 'Unit', 'Product Description', 'Note 1', 'Note 2', 'Image Count']
    ws.append(columns)

    product = Product.objects.prefetch_related('pairing_set', 'images').order_by('-id')

    for prod in product:
        pairing_set_values = ', '.join([ps.pair_value for ps in prod.pairing_set.all()]) if prod.pairing_set.exists() else None
        ws.append([
            prod.parent_code,
            prod.child_code,
            prod.location,  
            prod.stock,
            prod.kpo,
            pairing_set_values,
            prod.weight,
            prod.thai_baht,
            prod.usd_rate,
            prod.euro_rate,
            (prod.tag.name if getattr(prod, 'tag', None) else ''),
            (prod.unit or ''),
            (prod.description or ''),
            prod.note_1,
            prod.note_2,
            prod.images.count(),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=karen_report.xlsx'
    wb.save(response)

    return response


def export_selected_to_excel(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('ids[]')  
        if not selected_ids:
            return JsonResponse({'success': False, 'error': 'No products selected.'}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Selected Products'

        columns = ['Parent Code', 'Child Code', 'Location', 'QTY', 'kpo', 'pairing_set', 'weight', 'thai_baht', 'usd_rate', 'euro_rate', 'Category', 'Unit', 'Product Description', 'Note 1', 'Note 2', 'Image Count']
        ws.append(columns)

        products = Product.objects.filter(id__in=selected_ids).prefetch_related('pairing_set', 'images')

        for prod in products:
            pairing_set_values = ', '.join([ps.pair_value for ps in prod.pairing_set.all()]) if prod.pairing_set.exists() else None
            ws.append([
                prod.parent_code,
                prod.child_code,
                prod.location,
                prod.stock,
                prod.kpo,
                pairing_set_values,
                prod.weight,
                prod.thai_baht,
                prod.usd_rate,
                prod.euro_rate,
                (prod.tag.name if getattr(prod, 'tag', None) else ''),
                (prod.unit or ''),
                (prod.description or ''),
                prod.note_1,
                prod.note_2,
                prod.images.count(),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=selected_products.xlsx'
        wb.save(response)

        return response

    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)


@login_required
def pairing_set_api(request):
    pair_objs = PairingSet.objects.all().order_by('-id')
    pair_objs = ProductFilter(request.GET, queryset=pair_objs)
    data = [{'id': pair_obj.id, 'name': pair_obj.pair_value} for pair_obj in pair_objs.qs]

    data = custom_paginate_queryset(
        queryset=data,
        page=int(request.GET.get('page', 1)),
        page_size=int(request.GET.get('page_size', 10))
    )
    return JsonResponse(data, safe=False)


@csrf_exempt
def image_api(request):
    """API endpoint for image management operations"""
    if request.method == 'GET':
        # Get all images with pagination
        images = Image.objects.prefetch_related('product_set').all().order_by('-id')
        
        # Apply search filter if provided
        search_query = request.GET.get('search', '').strip()
        if search_query:
            images = images.filter(image__icontains=search_query)
        
        # Apply filter for linked/unlinked images
        filter_type = request.GET.get('filter', '').strip()
        if filter_type == 'linked':
            images = images.filter(product__isnull=False).distinct()
        elif filter_type == 'unlinked':
            images = images.filter(product__isnull=True)
        
        # Prepare image data
        image_data = []
        for image in images:
            linked_products = image.product_set.all()
            image_data.append({
                'id': image.id,
                'alt_text': image.image.name.split('/')[-1] if image.image else 'No name',
                'image_url': image.image.url if image.image else '',
                'uploaded_at': image.id,  # Using id as placeholder for upload date
                'linked_products_count': linked_products.count(),
                'linked_products': [{
                    'id': p.id,
                    'name': f"{p.parent_code} - {p.child_code}",
                    'parent_code': p.parent_code,
                    'child_code': p.child_code
                } for p in linked_products[:5]]  # Limit to first 5 for performance
            })
        
        # Apply pagination
        paginated_data = custom_paginate_queryset(
            queryset=image_data,
            page=int(request.GET.get('page', 1)),
            page_size=int(request.GET.get('page_size', 20))
        )
        
        return JsonResponse(paginated_data, safe=False)
    
    elif request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'link_products':
            image_id = request.POST.get('image_id')
            product_ids = request.POST.get('product_ids', '').split(',')
            
            try:
                image = Image.objects.get(id=image_id)
                selected_product_ids = [int(pid) for pid in product_ids if pid.strip()]
                
                # Get currently linked products
                current_linked_products = set(image.product_set.values_list('id', flat=True))
                selected_products_set = set(selected_product_ids)
                
                # Add new links (products that are selected but not currently linked)
                products_to_add = selected_products_set - current_linked_products
                for product_id in products_to_add:
                    try:
                        product = Product.objects.get(id=product_id)
                        product.images.add(image)
                        
                        # Create persistent link
                        ProductImageLink.objects.get_or_create(
                            image=image,
                            parent_code=product.parent_code,
                            child_code=product.child_code
                        )
                    except Product.DoesNotExist:
                        continue
                
                # Remove links (products that were linked but are no longer selected)
                products_to_remove = current_linked_products - selected_products_set
                for product_id in products_to_remove:
                    try:
                        product = Product.objects.get(id=product_id)
                        product.images.remove(image)
                        
                        # Remove persistent link when manually unlinked
                        ProductImageLink.objects.filter(
                            image=image,
                            parent_code=product.parent_code,
                            child_code=product.child_code
                        ).delete()
                    except Product.DoesNotExist:
                        continue
                
                return JsonResponse({'success': True, 'message': 'Products linked successfully'})
            except Image.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Image not found'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        
        elif action == 'auto_link':
            try:
                linked_count = 0
                images = Image.objects.all()
                
                for image in images:
                    if image.image:
                        # Extract filename without extension
                        filename = image.image.name.split('/')[-1].split('.')[0]
                        
                        # Try to match with product codes
                        matching_products = Product.objects.filter(
                            models.Q(parent_code__icontains=filename) |
                            models.Q(child_code__icontains=filename)
                        )
                        
                        if matching_products.exists():
                            for product in matching_products:
                                if not product.images.filter(id=image.id).exists():
                                    product.images.add(image)
                                    
                                    # Create persistent link
                                    ProductImageLink.objects.get_or_create(
                                        image=image,
                                        parent_code=product.parent_code,
                                        child_code=product.child_code
                                    )
                            linked_count += 1
                
                return JsonResponse({
                    'success': True, 
                    'message': f'Auto-linking completed successfully',
                    'linked_count': linked_count
                })
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        
        elif action == 'delete_image':
            image_id = request.POST.get('image_id')
            
            try:
                image = Image.objects.get(id=image_id)
                
                # Store the file path before deleting the object
                image_path = None
                if image.image:
                    image_path = image.image.path
                
                # Remove the image from all products first
                for product in image.product_set.all():
                    product.images.remove(image)
                
                # Delete the image object from database
                image.delete()
                
                # Delete the physical file from media folder
                if image_path and os.path.exists(image_path):
                    os.remove(image_path)
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Image deleted successfully'
                })
            except Image.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Image not found'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': f'Error deleting image: {str(e)}'})
        
        elif action == 'bulk_delete_images':
            image_ids_str = request.POST.get('image_ids')
            
            if not image_ids_str:
                return JsonResponse({'success': False, 'message': 'No image IDs provided'})
            
            try:
                # Parse the comma-separated image IDs
                image_ids = [int(id.strip()) for id in image_ids_str.split(',') if id.strip()]
                
                if not image_ids:
                    return JsonResponse({'success': False, 'message': 'No valid image IDs provided'})
                
                # Get all images to delete
                images_to_delete = Image.objects.filter(id__in=image_ids)
                
                if not images_to_delete.exists():
                    return JsonResponse({'success': False, 'message': 'No images found with the provided IDs'})
                
                deleted_count = 0
                failed_deletions = []
                
                for image in images_to_delete:
                    try:
                        # Store the file path before deleting the object
                        image_path = None
                        if image.image:
                            image_path = image.image.path
                        
                        # Remove the image from all products first
                        for product in image.product_set.all():
                            product.images.remove(image)
                        
                        # Delete the image object from database
                        image.delete()
                        
                        # Delete the physical file from media folder
                        if image_path and os.path.exists(image_path):
                            os.remove(image_path)
                        
                        deleted_count += 1
                        
                    except Exception as e:
                        failed_deletions.append(f"Image ID {image.id}: {str(e)}")
                
                if failed_deletions:
                    return JsonResponse({
                        'success': True,
                        'deleted_count': deleted_count,
                        'message': f'Deleted {deleted_count} images. Failed to delete some images: {"; ".join(failed_deletions)}'
                    })
                else:
                    return JsonResponse({
                        'success': True,
                        'deleted_count': deleted_count,
                        'message': f'Successfully deleted {deleted_count} images'
                    })
                    
            except ValueError:
                return JsonResponse({'success': False, 'message': 'Invalid image ID format'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': f'Error during bulk deletion: {str(e)}'})

        elif action == 'delete_all_images':
            try:
                images = Image.objects.all()
                image_paths = []
                for image in images:
                    try:
                        if image.image and hasattr(image.image, 'path'):
                            image_paths.append(image.image.path)
                    except Exception:
                        continue

                deleted_count = images.count()
                images.delete()

                for image_path in image_paths:
                    try:
                        if image_path and os.path.exists(image_path):
                            os.remove(image_path)
                    except Exception:
                        continue

                return JsonResponse({
                    'success': True,
                    'deleted_count': deleted_count,
                    'message': f'Successfully deleted {deleted_count} images'
                })
            except Exception as e:
                return JsonResponse({'success': False, 'message': f'Error deleting all images: {str(e)}'})
        
        return JsonResponse({'success': False, 'message': 'Invalid action'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@csrf_exempt
@login_required
def product_images_api(request, product_id):
    """API endpoint to get all images for a specific product"""
    if request.method == 'GET':
        try:
            product = Product.objects.get(id=product_id)
            images = product.image.all() if hasattr(product, 'image') else []
            
            image_data = []
            for image in images:
                if image.image:
                    image_data.append({
                        'url': image.image.url,
                        'alt': f'{product.parent_code} - {product.child_code}'
                    })
            
            # If no images found, return the main product image if available
            if not image_data and product.images.exists():
                main_image = product.images.first()
                image_data.append({
                    'url': main_image.image.url,
                    'alt': f'{product.parent_code} - {product.child_code}'
                })
            
            return JsonResponse({
                'success': True,
                'images': image_data,
                'product_name': f'{product.parent_code} - {product.child_code}',
                'image_count': len(image_data)
            })
            
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Product not found'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def pairing_set_view(request):
    if request.method == 'POST':
        data = request.POST
        method_type = data.get('method_type')
        if method_type == 'export':
            # Export pairing sets to Excel
            pairing_sets = PairingSet.objects.all().order_by('pair_value')
            
            # Create workbook and worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Pairing Sets"
            
            # Add headers
            headers = ['ID', 'Pair Value']
            for col_num, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col_num, value=header)
            
            # Add data
            for row_num, pairing_set in enumerate(pairing_sets, 2):
                worksheet.cell(row=row_num, column=1, value=pairing_set.id)
                worksheet.cell(row=row_num, column=2, value=pairing_set.pair_value)
            
            # Create HTTP response with Excel file
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=pairing_sets_export.xlsx'
            workbook.save(response)
            return response
        elif method_type == 'create':
            # Check if Excel file is uploaded
            if 'excel_file' in request.FILES:
                excel_file = request.FILES['excel_file']
                
                # Validate file extension
                if not excel_file.name.endswith(('.xlsx', '.xls')):
                    messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
                    return redirect('pairing_set')
                
                try:
                    # Load the Excel file
                    workbook = load_workbook(excel_file)
                    sheet = workbook.active
                    
                    pair_values = []
                    # Read values from the first column, starting from row 1
                    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
                        if row[0] is not None:
                            value = str(row[0]).strip()
                            if value:  # Only add non-empty values
                                pair_values.append(value)
                    
                    if not pair_values:
                        messages.error(request, 'No valid pairing set values found in the Excel file.')
                        return redirect('pairing_set')
                    
                    # Process the values
                    created_count = 0
                    skipped_count = 0
                    error_values = []
                    
                    for single_pair_value in pair_values:
                        if not single_pair_value:
                            continue
                        
                        # Check if pairing set already exists
                        if PairingSet.objects.filter(pair_value=single_pair_value).exists():
                            skipped_count += 1
                            continue
                        
                        try:
                            PairingSet.objects.create(pair_value=single_pair_value)
                            created_count += 1
                        except Exception as e:
                            error_values.append(f"{single_pair_value}: {str(e)}")
                    
                    # Prepare success message
                    success_msg = f'Excel upload completed: {created_count} pairing set(s) created'
                    if skipped_count > 0:
                        success_msg += f', {skipped_count} skipped (already exist)'
                    
                    if error_values:
                        success_msg += f'. Errors: {"; ".join(error_values[:3])}'  # Show first 3 errors
                        if len(error_values) > 3:
                            success_msg += f' and {len(error_values) - 3} more errors.'
                    
                    if created_count > 0:
                        messages.success(request, success_msg)
                    else:
                        messages.warning(request, success_msg)
                        
                except Exception as e:
                    messages.error(request, f'Error processing Excel file: {str(e)}')
                
                return redirect('pairing_set')
            else:
                messages.error(request, 'Please upload an Excel file.')
                return redirect('pairing_set')
        elif method_type == 'delete':
            pair_id = data.get('product_id')
            if not pair_id:
                messages.error(request, 'Invalid pair ID.')
                return redirect('pairing_set')
            pair = PairingSet.objects.get(id=pair_id)
            pair.delete()
            messages.success(request, 'Successfully deleted pairing set.')
            return redirect('pairing_set')
        elif method_type == 'delete_multiple':
            selected_ids = request.POST.getlist('selected_pairs')
            if not selected_ids:
                messages.error(request, 'No pairing sets selected for deletion.')
                return redirect('pairing_set')
            try:
                deleted_count = PairingSet.objects.filter(id__in=selected_ids).count()
                PairingSet.objects.filter(id__in=selected_ids).delete()
                messages.success(request, f'Successfully deleted {deleted_count} pairing set(s).')
            except Exception as e:
                messages.error(request, f'Error deleting pairing sets: {str(e)}')
            return redirect('pairing_set')
    return render(request, 'pairing_set.html')

@login_required
def pairing_set_print_view(request):
    """Print view for pairing sets"""
    pairing_sets = PairingSet.objects.all().order_by('pair_value')
    context = {
        'pairing_sets': pairing_sets,
        'total_count': pairing_sets.count()
    }
    return render(request, 'pairing_set_print.html', context)


@csrf_exempt
@login_required()
def get_unlinked_images(request):
    """API endpoint to get images that are not linked to any products"""
    if request.method == 'GET':
        # Get images that have no products linked
        unlinked_images = Image.objects.filter(product__isnull=True).distinct()
        
        image_data = []
        for image in unlinked_images:
            image_data.append({
                'id': image.id,
                'name': image.image.name.split('/')[-1] if image.image else 'No name',
                'url': image.image.url if image.image else ''
            })
        
        return JsonResponse({
            'success': True,
            'images': image_data,
            'count': len(image_data)
        })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@csrf_exempt
@login_required()
def search_products_for_linking(request):
    """API endpoint to search products for image linking"""
    if request.method == 'GET':
        query = request.GET.get('q', '').strip()
        
        if not query:
            return JsonResponse({'success': False, 'message': 'No search query provided'})
        
        # Search products by parent_code or child_code
        products = Product.objects.filter(
            Q(parent_code__icontains=query) | Q(child_code__icontains=query)
        ).order_by('child_code')[:20]  # Limit to 20 results
        
        product_data = []
        for product in products:
            product_data.append({
                'id': product.id,
                'parent_code': product.parent_code,
                'child_code': product.child_code,
                'location': product.location,
                'display_name': f"{product.parent_code} - {product.child_code}"
            })
        
        return JsonResponse({
            'success': True,
            'products': product_data,
            'count': len(product_data)
        })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})



# Product.objects.all().delete()
# print('delete')
# print(Image.objects.count())
# a = Product.objects.all()
# for i in a:
#     if i.images_names:
#         for img in i.images_names.all():
#             print(i.id, img.name)

# lis = ['BS0598-6.5', 'ES2314', 'NS0429ME-18', 'ES2313', 'ES2311', 'ES1652', 'ES2312', 'PS0789', 'BS0595ME-6.5', 'NS0429ME-18', 'NS0427ME-18', 'RS0425MPSE-6', 'RS0425MPSE-7', 'RS0425MPSE-8', 'RS0425MPSE-9']
# for i in lis:
#     check_obj = PairingSet.objects.filter(pair_value=i).first()
#     if not check_obj:
#         PairingSet.objects.create(pair_value=i)
#         print('save')

# def generate_QR_code(request):
#     if request.method == 'POST':
#         data = request.POST['text']
#         img = make(data)
#         img_name = f'{str(time())}.png'
#         img.save(settings.MEDIA_ROOT + '/' + img_name)
#         context = {
#             'img_name': img_name
#         }
#         return render(request, 'index.html', context)
#     return render(request, 'index.html')
# for i in Product.objects.all():
#     if i.qrcode_image and i.barcode_image:
#         print(i.id)




# import os
# from datetime import datetime
# import shutil

# def db_backup():
#     db_path = settings.DATABASES['default']['NAME']
#     backup_dir = os.path.join(settings.BASE_DIR, 'db_backup_files') 
#     os.makedirs(backup_dir, exist_ok=True)
    
#     timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
#     backup_file = os.path.join(backup_dir, f'db_backup_{timestamp}.sqlite3')
    
#     try:
#         shutil.copy2(db_path, backup_file)
#     except Exception as e:
#         print(f'Error during backup: {e}')

# db_backup()
# PairingSet.objects.all().delete()
# ImageName.objects.all().delete()
# Product.objects.all().delete()


# Utility functions for image management
def auto_link_images_to_products():
    """Utility function to automatically link images to products based on naming patterns"""
    unlinked_images = Image.objects.filter(product__isnull=True)
    linked_count = 0
    
    for image in unlinked_images:
        if image.image:
            # Extract filename without extension
            filename = image.image.name.split('/')[-1].split('.')[0]
            
            # Try to find matching products
            matching_products = Product.objects.filter(
                Q(parent_code__icontains=filename) | Q(child_code__icontains=filename)
            )
            
            for product in matching_products:
                if not product.images.filter(id=image.id).exists():
                    product.images.add(image)
                    
                    # Create persistent link
                    ProductImageLink.objects.get_or_create(
                        image=image,
                        parent_code=product.parent_code,
                        child_code=product.child_code
                    )
                    linked_count += 1
    
    return linked_count

# Cart Management Views
@login_required
def cart_management_view(request):
    """Main cart management view with customer list"""
    customers = Customer.objects.all().order_by('name')
    
    # Handle customer creation
    if request.method == 'POST' and request.POST.get('action') == 'create_customer':
        customer_name = request.POST.get('customer_name', '').strip()
        if customer_name:
            customer = Customer.objects.create(name=customer_name)
            messages.success(request, f'Customer "{customer_name}" created successfully!')
            return redirect('cart_management')
        else:
            messages.error(request, 'Customer name is required!')
    
    context = {
        'customers': customers,
    }
    return render(request, 'cart_management.html', context)

@csrf_exempt
@login_required
def customer_api(request):
    """API for customer CRUD operations"""
    if request.method == 'GET':
        customers = Customer.objects.all().order_by('name')
        customer_list = []
        for customer in customers:
            active_cart = customer.carts.filter(is_active=True).first()
            customer_list.append({
                'id': customer.id,
                'name': customer.name,
                'locked': customer.locked,
                'created_at': customer.created_at.strftime('%Y-%m-%d %H:%M'),
                'cart_count': customer.carts.count(),
                'active_cart_id': active_cart.id if active_cart else None,
                'active_cart_items': active_cart.get_total_items() if active_cart else 0,
            })
        return JsonResponse({'customers': customer_list})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            if action == 'create':
                name = data.get('name', '').strip()
                if not name:
                    return JsonResponse({'success': False, 'message': 'Customer name is required'})
                
                customer = Customer.objects.create(name=name)
                return JsonResponse({
                    'success': True, 
                    'message': f'Customer "{name}" created successfully!',
                    'customer': {
                        'id': customer.id,
                        'name': customer.name,
                        'created_at': customer.created_at.strftime('%Y-%m-%d %H:%M'),
                        'cart_count': 0,
                        'active_cart_id': None,
                        'active_cart_items': 0,
                    }
                })
            
            elif action == 'update':
                customer_id = data.get('id')
                name = data.get('name', '').strip()
                
                if not customer_id or not name:
                    return JsonResponse({'success': False, 'message': 'Customer ID and name are required'})
                
                customer = get_object_or_404(Customer, id=customer_id)
                customer.name = name
                customer.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Customer updated successfully!',
                    'customer': {
                        'id': customer.id,
                        'name': customer.name,
                        'created_at': customer.created_at.strftime('%Y-%m-%d %H:%M'),
                    }
                })
            
            elif action == 'delete':
                customer_id = data.get('id')
                if not customer_id:
                    return JsonResponse({'success': False, 'message': 'Customer ID is required'})
                
                customer = get_object_or_404(Customer, id=customer_id)
                customer_name = customer.name
                customer.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Customer "{customer_name}" deleted successfully!'
                })
            
            else:
                return JsonResponse({'success': False, 'message': 'Invalid action'})
                
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})


@login_required
def export_customer_cart_excel(request, customer_id):
    """Export customer cart to Excel file"""
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        cart, created = Cart.objects.get_or_create(customer=customer, is_active=True)
        cart_items = CartItem.objects.filter(cart=cart).select_related('product')
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f'Cart - {customer.name}'
        
        from openpyxl.styles import Font, Alignment, Border, Side
        thin = Side(style='thin', color='000000')
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ci_title = ws.cell(row=1, column=1, value='CUSTOMER INFO')
        ci_title.font = Font(bold=True)
        ws.cell(row=2, column=1, value='Name:')
        ws.cell(row=2, column=2, value=customer.name or '')
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)
        ws.cell(row=3, column=1, value='Company Name:')
        ws.cell(row=3, column=2, value=(getattr(customer, 'company_name', '') or ''))
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
        ws.cell(row=4, column=1, value='Address:')
        address_display = (cart.address_override or getattr(customer, 'address', '') or '')
        ws.cell(row=4, column=2, value=address_display)
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=4)
        ws.cell(row=5, column=1, value='Phone:')
        ws.cell(row=5, column=2, value=(getattr(customer, 'phone', '') or ''))
        ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)
        ws.cell(row=6, column=1, value='Email:')
        ws.cell(row=6, column=2, value=(getattr(customer, 'email', '') or ''))
        ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=4)
        if cart.notes:
            ws.cell(row=7, column=1, value='Notes:')
            ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=4)
            ws.cell(row=7, column=2, value=str(cart.notes))

        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=8)
        di_title = ws.cell(row=1, column=5, value='DOCUMENT INFO')
        di_title.font = Font(bold=True)
        ws.cell(row=2, column=5, value='Date:')
        ws.cell(row=2, column=6, value=timezone.now().strftime('%Y-%m-%d %H:%M'))
        ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=8)
        salesperson_name = (getattr(cart, 'sales_person', None) or getattr(request.user, 'get_full_name', lambda: '')() or getattr(request.user, 'username', ''))
        ws.cell(row=3, column=5, value='Salesperson:')
        ws.cell(row=3, column=6, value=salesperson_name)
        ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=8)
        ws.cell(row=4, column=5, value='Doc. Ref.:')
        ws.cell(row=4, column=6, value=str(getattr(cart, 'doc_ref', None) or getattr(cart, 'id', '')))
        ws.merge_cells(start_row=4, start_column=6, end_row=4, end_column=8)
        ws.cell(row=5, column=5, value='Customer Code:')
        ws.cell(row=5, column=6, value=str(getattr(cart, 'customer_code', None) or getattr(customer, 'id', '')))
        ws.merge_cells(start_row=5, start_column=6, end_row=5, end_column=8)

        for r in range(1, 8):
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = border
        for r in range(1, 8):
            for c in range(5, 9):
                ws.cell(row=r, column=c).border = border

        ws.append([])
        
        currency = (request.GET.get('currency') or 'THB').upper()
        if currency not in ('THB', 'USD', 'EUR'):
            currency = 'THB'
        
        # Parse selected columns from query params
        selected_cols_param = request.GET.get('cols', '')
        selected_cols = [c.strip() for c in selected_cols_param.split(',') if c.strip()] or [
            'product_code', 'location', 'qty', 'wt_g', 'price_thb', 'amount_thb', 'usd', 'euro', 'note1', 'note2'
        ]

        def to_float(val):
            try:
                return float(val)
            except Exception:
                try:
                    return float(str(val).strip())
                except Exception:
                    return 0.0

        headers = []
        for c in selected_cols:
            if c == 'product_code': headers.append('Product Code')
            elif c == 'picture': headers.append('Picture')
            elif c == 'name': headers.append('Product Name')
            elif c == 'wt_g': headers.append('WT. (g)')
            elif c == 'qty': headers.append('Quantity')
            elif c == 'price_thb': headers.append(f'Unit Price ({currency})')
            elif c == 'amount_thb': headers.append(f'Total Price ({currency})')
            elif c == 'location': headers.append('Location')
            elif c == 'kpo': headers.append('KPO')
            elif c == 'pairing_set': headers.append('Pairing Set')
            elif c == 'note1': headers.append('Note 1')
            elif c == 'note2': headers.append('Note 2')
            elif c == 'thb': headers.append('THB')
            elif c == 'usd': headers.append('USD')
            elif c == 'euro': headers.append('EURO')
        product_header_row = ws.max_row + 1
        ws.append(headers)

        items_for_export = []
        total_amount = 0.0
        total_quantity = 0
        total_weight = 0.0
        earring_count = 0
        ring_count = 0
        bracelet_bangle_count = 0
        necklace_count = 0
        others_count = 0
        for ci in cart_items:
            p = ci.product
            if currency == 'USD':
                unit = to_float(p.usd_rate or 0)
                label = 'USD'
            elif currency == 'EUR':
                unit = to_float(p.euro_rate or 0)
                label = 'EUR'
            else:
                unit = to_float(p.thai_baht or 0)
                label = 'THB'
            amount = unit * ci.quantity if unit else 0.0
            total_amount += amount
            total_quantity += ci.quantity
            total_weight += to_float(p.weight or 0) * ci.quantity
            tag_name = (p.tag.name if p.tag else '') or ''
            t = tag_name.lower()
            if 'earring' in t:
                earring_count += ci.quantity
            elif 'ring' in t:
                ring_count += ci.quantity
            elif 'bracelet' in t or 'bangle' in t:
                bracelet_bangle_count += ci.quantity
            elif 'necklace' in t:
                necklace_count += ci.quantity
            else:
                others_count += ci.quantity
            items_for_export.append({
                'code': f"{p.child_code}",
                'image_url': '',
                'name': tag_name if tag_name else '-',
                'weight': to_float(p.weight or 0),
                'quantity': ci.quantity,
                'unit': unit,
                'amount': amount,
                'location': p.location or '-',
                'kpo': p.kpo or '-',
                'pairing_set': ', '.join([ps.pair_value for ps in p.pairing_set.all()]) if hasattr(p, 'pairing_set') else '-',
                'note1': p.note_1 or '-',
                'note2': p.note_2 or '-',
                'thb': to_float(p.thai_baht or 0),
                'usd': to_float(p.usd_rate or 0),
                'eur': to_float(p.euro_rate or 0),
            })

        for it in items_for_export:
            row = []
            for c in selected_cols:
                if c == 'product_code': row.append(it['code'])
                elif c == 'picture': row.append(it['image_url'])
                elif c == 'name': row.append(it['name'])
                elif c == 'wt_g': row.append(it['weight'])
                elif c == 'qty': row.append(it['quantity'])
                elif c == 'price_thb': row.append(it['unit'])
                elif c == 'amount_thb': row.append(it['amount'])
                elif c == 'location': row.append(it['location'])
                elif c == 'kpo': row.append(it['kpo'])
                elif c == 'pairing_set': row.append(it['pairing_set'])
                elif c == 'note1': row.append(it['note1'])
                elif c == 'note2': row.append(it['note2'])
                elif c == 'thb': row.append(it['thb'])
                elif c == 'usd': row.append(it['usd'])
                elif c == 'euro': row.append(it['eur'])
            ws.append(row)

        ws.append([])
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        label_font = Font(bold=True)
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        thin = Side(style='thin', color='000000')
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        start_row = ws.max_row + 1
        shipping_amount = float(cart.shipping_amount or 0)
        deposit_amount = float(cart.deposit_amount or 0)
        grand_total = (total_amount + shipping_amount) - deposit_amount
        gross_weight_val = float(cart.gross_weight or 0)
        ws.append(["Total", f"{currency} {total_amount:.2f}"])
        ws.append(["Shipping", f"{currency} {shipping_amount:.2f}"])
        ws.append(["Deposit", f"{currency} {deposit_amount:.2f}"])
        ws.append(["Grand Total", f"{currency} {grand_total:.2f}"])
        for r in range(start_row, start_row + 4):
            ws[f"A{r}"] .fill = black_fill
            ws[f"A{r}"] .font = white_font
            ws[f"B{r}"] .fill = black_fill
            ws[f"B{r}"] .font = white_font
            ws[f"A{r}"] .border = border
            ws[f"B{r}"] .border = border

        ws.append([])
        ws.append(["SUMMARY"]) 
        ws.cell(row=ws.max_row, column=1).fill = black_fill
        ws.cell(row=ws.max_row, column=1).font = white_font
        ws.append([f"Total Items: {len(items_for_export)} Pcs."])
        ws.append([f"Total Quantity: {total_quantity} Pcs."])
        ws.append([f"Total Net Weight: {total_weight:.2f} g."])
        ws.append([f"Total Gross Weight: {gross_weight_val:.2f} g."])
        ws.append(["Category"]) 
        ws.append([f"Earring: {earring_count} Prs."])
        ws.append([f"Ring: {ring_count} Pcs."])
        ws.append([f"Bracelet and Bangle: {bracelet_bangle_count} Pcs."])
        ws.append([f"Necklace: {necklace_count} Pcs."])
        ws.append([f"Other Accessories (Ex. Beads): {others_count} Pcs."])
        ws.append([f"Notes: {str(cart.notes) if cart.notes else ''}"])
        
        # Style the worksheet
        from openpyxl.styles import Font, Alignment, PatternFill
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        for row in ws.iter_rows(min_row=product_header_row, max_row=product_header_row):
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths (safe for merged cells)
        from openpyxl.utils import get_column_letter
        for i, column_cells in enumerate(
            ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=product_header_row, max_row=ws.max_row), start=1
        ):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value is None:
                        continue
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="cart_{customer.name}_{customer.id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        )
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting cart: {str(e)}')
        return redirect('customer_cart', customer_id=customer_id)


@login_required
def print_customer_cart(request, customer_id):
    """Render customer cart for printing"""
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        cart, created = Cart.objects.get_or_create(customer=customer, is_active=True)
        cart_items = CartItem.objects.filter(cart=cart).select_related('product')
        currency = (request.GET.get('currency') or 'THB').upper()
        if currency not in ('THB', 'USD', 'EUR'):
            currency = 'THB'
        
        # Parse selected columns
        selected_cols_param = request.GET.get('cols', '')
        selected_columns = [c.strip() for c in selected_cols_param.split(',') if c.strip()] or [
            'product_code','qty','wt_g','price_thb','amount_thb','location','note1','note2','usd','euro'
        ]

        # Helpers
        def to_float(val):
            d = _parse_price_to_decimal(val)
            if d is not None:
                try:
                    return float(d)
                except Exception:
                    return 0.0
            try:
                return float(val)
            except Exception:
                try:
                    return float(str(val).strip())
                except Exception:
                    return 0.0

        # Build items for print in selected currency
        items_for_print = []
        total_amount = 0.0
        for ci in cart_items:
            p = ci.product
            if currency == 'USD':
                unit = to_float(p.usd_rate or 0)
                label = 'USD'
            elif currency == 'EUR':
                unit = to_float(p.euro_rate or 0)
                label = 'EUR'
            else:
                unit = to_float(p.thai_baht or 0)
                label = 'THB'
            amount = unit * ci.quantity if unit else 0.0
            total_amount += amount
            # extra fields for dynamic columns
            try:
                image_url = ''
                if p.images.exists():
                    first_image = p.images.first()
                    image_url = request.build_absolute_uri(first_image.image.url)
            except Exception:
                image_url = ''
            pairing_values = []
            try:
                pairing_values = [ps.pair_value for ps in p.pairing_set.all()]
            except Exception:
                pairing_values = []
            items_for_print.append({
                'code': f"{p.child_code}",
                'name': p.tag.name if p.tag else '-',
                'location': p.location or '-',
                'quantity': ci.quantity,
                'unit': unit,
                'amount': amount,
                'currency_label': label,
                'image_url': image_url,
                'weight': to_float(p.weight or 0),
                'kpo': p.kpo or '-',
                'note1': p.note_1 or '-',
                'note2': p.note_2 or '-',
                'pairing_set': ", ".join(pairing_values) if pairing_values else '-',
                'thb': to_float(p.thai_baht or 0),
                'usd': to_float(p.usd_rate or 0),
                'eur': to_float(p.euro_rate or 0),
            })

        # Calculate totals and annotate items with computed fields
        total_quantity = sum(item.quantity for item in cart_items)
        total_weight = sum(
            float(item.product.weight or 0) * item.quantity 
            for item in cart_items
        )
        for item in cart_items:
            try:
                item.amount_thb = to_float(item.product.thai_baht) * item.quantity
            except Exception:
                item.amount_thb = 0.0
        shipping = float(cart.shipping_amount or 0)
        deposit = float(cart.deposit_amount or 0)
        grand_total = total_amount + shipping - deposit

        # Category summary breakdown to match app print layout
        earring_count = 0
        ring_count = 0
        bracelet_bangle_count = 0
        necklace_count = 0
        others_count = 0
        for item in cart_items:
            tag_name = (item.product.tag.name if item.product.tag else '') or ''
            t = tag_name.lower()
            if 'earring' in t:
                earring_count += item.quantity
            elif 'ring' in t:
                ring_count += item.quantity
            elif 'bracelet' in t or 'bangle' in t:
                bracelet_bangle_count += item.quantity
            elif 'necklace' in t:
                necklace_count += item.quantity
            else:
                others_count += item.quantity
        
        context = {
            'customer': customer,
            'cart': cart,
            'items_for_print': items_for_print,
            'total_items': cart_items.count(),
            'total_quantity': total_quantity,
            'total_weight': total_weight,
            'total_amount': total_amount,
            'shipping': shipping,
            'deposit': deposit,
            'grand_total': grand_total,
            'print_date': timezone.now(),
            'selected_columns': selected_columns,
            'earring_count': earring_count,
            'ring_count': ring_count,
            'bracelet_bangle_count': bracelet_bangle_count,
            'necklace_count': necklace_count,
            'others_count': others_count,
            'currency_label': currency,
            'salesperson_name': (getattr(cart, 'sales_person', None) or getattr(request.user, 'get_full_name', lambda: '')() or getattr(request.user, 'username', '')),
            'doc_ref': str(getattr(cart, 'doc_ref', None) or getattr(cart, 'id', '')),
            'customer_code': str(getattr(cart, 'customer_code', None) or getattr(customer, 'id', '')),
            'address_display': (cart.address_override or getattr(customer, 'address', '') or ''),
        }
        
        return render(request, 'customer_cart_print.html', context)
        
    except Exception as e:
        messages.error(request, f'Error preparing cart for printing: {str(e)}')
        return redirect('customer_cart', customer_id=customer_id)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def customer_cart_view(request, customer_id):
    """View customer's cart details"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    # Get or create active cart for customer
    cart, created = Cart.objects.get_or_create(
        customer=customer,
        is_active=True,
        defaults={'customer': customer}
    )
    
    cart_items = cart.items.select_related('product').all()
    
    context = {
        'customer': customer,
        'cart': cart,
        'cart_items': cart_items,
        'products': Product.objects.prefetch_related('images').all().order_by('parent_code', 'child_code'),
    }
    return render(request, 'customer_cart.html', context)

@csrf_exempt
@login_required
def import_customer_cart_excel(request, customer_id):
    """Import products (code, quantity) from Excel into customer's active cart.
    Expected columns: Product Code, Quantity. Non-existent products are ignored.
    """
    customer = get_object_or_404(Customer, id=customer_id)
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'success': False, 'message': 'No file uploaded'})
        try:
            wb = load_workbook(filename=file, data_only=True)
            ws = wb.active

            # Detect header row and indices for code/quantity
            start_row = 1
            idx_code = 0
            idx_qty = 1
            first_row_values = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
            if first_row_values:
                header_lower = [str(v).strip().lower() if v is not None else '' for v in first_row_values]
                def find_index(options, default=None):
                    for opt in options:
                        if opt in header_lower:
                            return header_lower.index(opt)
                    return default
                code_idx = find_index(['product code', 'code'])
                qty_idx = find_index(['quantity', 'qty'])
                headers_present = (code_idx is not None) or (qty_idx is not None)
                if code_idx is not None:
                    idx_code = code_idx
                if qty_idx is not None:
                    idx_qty = qty_idx
                if headers_present:
                    start_row = 2

            cart, _ = Cart.objects.get_or_create(customer=customer, is_active=True)
            added = updated = ignored = processed = 0

            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not row:
                    continue
                code_cell = row[idx_code] if len(row) > idx_code else None
                qty_cell = row[idx_qty] if len(row) > idx_qty else 1
                if code_cell is None:
                    ignored += 1
                    continue
                code_str = str(code_cell).strip()
                try:
                    # Allow numeric cells like 2.0
                    qty = int(float(qty_cell)) if qty_cell is not None else 1
                except Exception:
                    qty = 1
                if not code_str or qty <= 0:
                    ignored += 1
                    continue

                # Find product by combined code or parent/child code
                product = None
                if '-' in code_str:
                    # Split on the LAST hyphen because parent_code may itself contain hyphens
                    try:
                        parent, child = code_str.rsplit('-', 1)
                        product = Product.objects.filter(parent_code=parent.strip(), child_code=child.strip()).first()
                    except ValueError:
                        product = None
                if not product:
                    product = Product.objects.filter(Q(child_code=code_str) | Q(parent_code=code_str)).first()
                if not product:
                    ignored += 1
                    continue

                cart_item, created = CartItem.objects.get_or_create(
                    cart=cart,
                    product=product,
                    defaults={'quantity': qty}
                )
                if created:
                    added += 1
                else:
                    cart_item.quantity += qty
                    cart_item.save()
                    updated += 1
                processed += 1

            return JsonResponse({
                'success': True,
                'message': f'Imported {processed} lines. Added: {added}, Updated: {updated}, Ignored: {ignored}.'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Import error: {str(e)}'})

    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@csrf_exempt
@login_required
def cart_api(request, customer_id):
    """API for cart operations"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            # Get or create active cart
            cart, created = Cart.objects.get_or_create(
                customer=customer,
                is_active=True,
                defaults={'customer': customer}
            )
            
            if action == 'add_item':
                product_id = data.get('product_id')
                quantity = int(data.get('quantity', 1))
                
                if not product_id:
                    return JsonResponse({'success': False, 'message': 'Product ID is required'})
                
                product = get_object_or_404(Product, id=product_id)
                
                cart_item, created = CartItem.objects.get_or_create(
                    cart=cart,
                    product=product,
                    defaults={'quantity': quantity}
                )
                
                if not created:
                    cart_item.quantity += quantity
                    cart_item.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Added {quantity}x {product.child_code} to cart'
                })
            
            elif action == 'update_item':
                item_id = data.get('item_id')
                quantity = int(data.get('quantity', 1))
                
                if not item_id:
                    return JsonResponse({'success': False, 'message': 'Item ID is required'})
                
                cart_item = get_object_or_404(CartItem, id=item_id, cart=cart)
                cart_item.quantity = quantity
                cart_item.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Cart item updated successfully'
                })
            
            elif action == 'remove_item':
                item_id = data.get('item_id')
                
                if not item_id:
                    return JsonResponse({'success': False, 'message': 'Item ID is required'})
                
                cart_item = get_object_or_404(CartItem, id=item_id, cart=cart)
                product_name = f"{cart_item.product.child_code}"
                cart_item.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Removed {product_name} from cart'
                })
            
            elif action == 'clear_cart':
                cart.items.all().delete()
                return JsonResponse({
                    'success': True,
                    'message': 'Cart cleared successfully'
                })
            
            elif action == 'update_cart_info':
                address_override = data.get('address_override')
                shipping_amount = data.get('shipping_amount')
                deposit_amount = data.get('deposit_amount')
                notes = data.get('notes')
                sales_person = data.get('sales_person')
                doc_ref = data.get('doc_ref')
                customer_code = data.get('customer_code')
                gross_weight = data.get('gross_weight')
                try:
                    cart.address_override = address_override
                    cart.shipping_amount = Decimal(str(shipping_amount or 0))
                    cart.deposit_amount = Decimal(str(deposit_amount or 0))
                    cart.notes = notes
                    cart.sales_person = sales_person
                    cart.doc_ref = doc_ref
                    cart.customer_code = customer_code
                    cart.gross_weight = Decimal(str(gross_weight or 0))
                    cart.save()
                    return JsonResponse({'success': True, 'message': 'Cart details updated'})
                except Exception as e:
                    return JsonResponse({'success': False, 'message': str(e)})
            else:
                return JsonResponse({'success': False, 'message': 'Invalid action'})
                
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    elif request.method == 'GET':
        # Get cart details
        cart = Cart.objects.filter(customer=customer, is_active=True).first()
        if not cart:
            return JsonResponse({'cart_items': []})
        
        currency = (request.GET.get('currency') or '').upper()
        if currency not in ('THB', 'USD', 'EUR'):
            currency = ''  # empty means auto-preferred fallback

        cart_items = []
        for item in cart.items.select_related('product').all():
            product = item.product
            # Determine price based on requested currency or fallback
            price_val = None
            price_raw = None
            currency_note = None
            if currency == 'THB':
                currency_note = 'THB'
                price_raw = product.thai_baht
                price_val = _parse_price_to_decimal(product.thai_baht)
            elif currency == 'USD':
                currency_note = 'USD'
                price_raw = product.usd_rate
                price_val = _parse_price_to_decimal(product.usd_rate)
            elif currency == 'EUR':
                currency_note = 'EUR'
                price_raw = product.euro_rate
                price_val = _parse_price_to_decimal(product.euro_rate)
            else:
                if product.thai_baht:
                    currency_note = 'THB'
                    price_raw = product.thai_baht
                    price_val = _parse_price_to_decimal(product.thai_baht)
                elif product.usd_rate:
                    currency_note = 'USD'
                    price_raw = product.usd_rate
                    price_val = _parse_price_to_decimal(product.usd_rate)
                elif product.euro_rate:
                    currency_note = 'EUR'
                    price_raw = product.euro_rate
                    price_val = _parse_price_to_decimal(product.euro_rate)

            amount_val = None
            if price_val is not None:
                try:
                    amount_val = price_val * Decimal(str(item.quantity))
                except Exception:
                    amount_val = None

            cart_items.append({
                'id': item.id,
                'product_id': product.id,
                'product_code': f"{product.child_code}",
                'product_location': product.location,
                'quantity': item.quantity,
                'price': float(price_val) if price_val is not None else None,
                'price_raw': (str(price_raw).strip() if price_raw is not None else None),
                'currency_note': currency_note,
                'amount': float(amount_val) if amount_val is not None else None,
                'added_at': item.added_at.strftime('%Y-%m-%d %H:%M'),
            })
        
        cart_info = {
            'address_override': cart.address_override,
            'shipping_amount': float(cart.shipping_amount or 0),
            'deposit_amount': float(cart.deposit_amount or 0),
            'notes': cart.notes,
            'sales_person': getattr(cart, 'sales_person', None),
            'doc_ref': getattr(cart, 'doc_ref', None),
            'customer_code': getattr(cart, 'customer_code', None),
            'gross_weight': float(getattr(cart, 'gross_weight', 0) or 0),
        }
        return JsonResponse({'cart_items': cart_items, 'cart_info': cart_info})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})


@csrf_exempt
def add_to_cart_api(request):
    """API endpoint for Android app to create customer and add product to cart"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        customer_name = data.get('customer_name', '').strip()
        product_code = data.get('product_code', '').strip()
        quantity = int(data.get('quantity', 1))
        
        # Validate input
        if not customer_name:
            return JsonResponse({'error': 'Customer name is required'}, status=400)
        
        if not product_code:
            return JsonResponse({'error': 'Product code is required'}, status=400)
        
        if quantity <= 0:
            return JsonResponse({'error': 'Quantity must be greater than 0'}, status=400)
        
        # Find the product by searching in both child_code and parent_code
        # This matches the logic used in single_product API
        product = Product.objects.filter(
            Q(child_code=product_code) | Q(parent_code=product_code)
        ).first()
        
        if not product:
            return JsonResponse({'error': 'Product not found'}, status=404)
        
        # Create or get customer
        customer, created = Customer.objects.get_or_create(
            name=customer_name,
            defaults={'created_at': timezone.now()}
        )
        
        # Get or create active cart for customer
        cart, cart_created = Cart.objects.get_or_create(
            customer=customer,
            is_active=True,
            defaults={'created_at': timezone.now()}
        )
        
        # Check if product already exists in cart
        cart_item, item_created = CartItem.objects.get_or_create(
            cart=cart,
            product=product,
            defaults={
                'quantity': quantity,
                'added_at': timezone.now()
            }
        )
        
        if not item_created:
            # Product already in cart, update quantity
            cart_item.quantity += quantity
            cart_item.save()
            message = f"Updated quantity for {product_code}. New quantity: {cart_item.quantity}"
        else:
            message = f"Added {quantity} x {product_code} to {customer_name}'s cart"
        
        return JsonResponse({
            'success': True,
            'message': message,
            'customer_id': customer.id,
            'customer_name': customer.name,
            'product_code': product_code,
            'quantity': cart_item.quantity,
            'cart_id': cart.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except ValueError as e:
        return JsonResponse({'error': f'Invalid data: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


@csrf_exempt
def customers_android_api(request):
    """API endpoint for Android app to get customer data without authentication"""
    if request.method == 'GET':
        customers = Customer.objects.all().order_by('name')
        customer_list = []
        for customer in customers:
            active_cart = customer.carts.filter(is_active=True).first()
            customer_list.append({
                'id': customer.id,
                'name': customer.name,
                'created_at': customer.created_at.strftime('%Y-%m-%d %H:%M'),
                'cart_count': customer.carts.count(),
                'active_cart_id': active_cart.id if active_cart else None,
                'active_cart_items': active_cart.get_total_items() if active_cart else 0,
            })
        return JsonResponse({'customers': customer_list})
    
    return JsonResponse({'error': 'Only GET method allowed'}, status=405)


@csrf_exempt
def customers_android_delete_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])
        if not isinstance(ids, list) or not ids:
            return JsonResponse({'error': 'ids must be a non-empty list'}, status=400)
        deleted_ids = []
        for cid in ids:
            try:
                customer = Customer.objects.get(id=cid)
                customer.delete()
                deleted_ids.append(cid)
            except Customer.DoesNotExist:
                pass
        return JsonResponse({'success': True, 'deleted_ids': deleted_ids, 'deleted_count': len(deleted_ids)})
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)


@csrf_exempt
def customers_android_create_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'error': 'Name is required'}, status=400)
        customer = Customer.objects.create(name=name)
        return JsonResponse({'success': True, 'customer': {'id': customer.id, 'name': customer.name}})
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)


@csrf_exempt
def customers_android_lock_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])
        lock = bool(data.get('lock', True))
        if not isinstance(ids, list) or not ids:
            return JsonResponse({'error': 'ids must be a non-empty list'}, status=400)
        updated = 0
        for cid in ids:
            try:
                customer = Customer.objects.get(id=cid)
                customer.locked = lock
                customer.save(update_fields=['locked'])
                updated += 1
            except Customer.DoesNotExist:
                continue
        return JsonResponse({'success': True, 'updated_count': updated, 'locked': lock})
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)


@csrf_exempt
def add_to_cart_bulk_android(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    try:
        data = json.loads(request.body)
        product_code = data.get('product_code', '').strip()
        quantity = int(data.get('quantity', 1))
        for_locked = bool(data.get('for_locked', True))
        if not product_code:
            return JsonResponse({'error': 'Product code is required'}, status=400)
        if quantity <= 0:
            return JsonResponse({'error': 'Quantity must be greater than 0'}, status=400)

        product = Product.objects.filter(Q(child_code=product_code) | Q(parent_code=product_code)).first()
        if not product:
            return JsonResponse({'error': 'Product not found'}, status=404)

        if for_locked:
            customers = Customer.objects.filter(locked=True)
        else:
            customer_ids = data.get('customer_ids', [])
            customers = Customer.objects.filter(id__in=customer_ids)

        added_count = 0
        for customer in customers:
            cart, _ = Cart.objects.get_or_create(customer=customer, is_active=True)
            item, created = CartItem.objects.get_or_create(cart=cart, product=product, defaults={'quantity': quantity})
            if not created:
                item.quantity += quantity
                item.save(update_fields=['quantity'])
            added_count += 1

        return JsonResponse({'success': True, 'affected_customers': added_count})
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)


@csrf_exempt
def cart_android_api(request, customer_id):
    """API endpoint for Android app to get and modify customer cart data without authentication"""
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        cart = customer.carts.filter(is_active=True).first()
        if not cart:
            cart = Cart.objects.create(customer=customer, is_active=True)

        if request.method == 'GET':
            cart_items = []
            for item in cart.items.all():
                image_url = ''
                if item.product.images.exists():
                    first_image = item.product.images.first()
                    image_url = request.build_absolute_uri(first_image.image.url)

                product_full = {
                    'id': item.product.id,
                    'image': item.product.images.first().image.url if item.product.images.exists() else '',
                    'parent_code': item.product.parent_code,
                    'child_code': item.product.child_code,
                    'location': item.product.location,
                    'stock': item.product.stock,
                    'kpo': item.product.kpo,
                    'weight': item.product.weight,
                    'thai_baht': item.product.thai_baht,
                    'thai_baht_value': (float(_parse_price_to_decimal(item.product.thai_baht)) if _parse_price_to_decimal(item.product.thai_baht) is not None else None),
                    'usd_rate': item.product.usd_rate,
                    'usd_rate_value': (float(_parse_price_to_decimal(item.product.usd_rate)) if _parse_price_to_decimal(item.product.usd_rate) is not None else None),
                    'euro_rate': item.product.euro_rate,
                    'euro_rate_value': (float(_parse_price_to_decimal(item.product.euro_rate)) if _parse_price_to_decimal(item.product.euro_rate) is not None else None),
                    'note_1': item.product.note_1,
                    'note_2': item.product.note_2,
                    'tag': item.product.tag.name if item.product.tag else '',
                    'pairing_set': [ps.id for ps in item.product.pairing_set.all()],
                    'image_count': item.product.images.count(),
                }

                price_dec = _parse_price_to_decimal(item.product.thai_baht)
                price_num = float(price_dec) if price_dec is not None else 0.0
                cart_items.append({
                    'id': item.id,
                    'product_id': item.product.id,
                    'product_code': f"{item.product.child_code}",
                    'product_name': f"{item.product.child_code}",
                    'product_location': item.product.location or '',
                    'price': price_num,
                    'price_raw': (str(item.product.thai_baht).strip() if item.product.thai_baht is not None else ''),
                    'image_url': image_url,
                    'quantity': item.quantity,
                    'added_at': item.added_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'product': product_full,
                })

            cart_info = {
                'address_override': cart.address_override,
                'shipping_amount': float(cart.shipping_amount or 0),
                'deposit_amount': float(cart.deposit_amount or 0),
                'notes': cart.notes,
                'sales_person': getattr(cart, 'sales_person', None),
                'doc_ref': getattr(cart, 'doc_ref', None),
                'customer_code': getattr(cart, 'customer_code', None),
                'gross_weight': float(getattr(cart, 'gross_weight', 0) or 0),
            }
            return JsonResponse({'cart_items': cart_items, 'cart_info': cart_info})

        elif request.method == 'POST':
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON data'}, status=400)

            action = data.get('action')
            if action == 'bulk_update':
                updates = data.get('updates', [])
                updated_count = 0
                for upd in updates:
                    item_id = upd.get('item_id')
                    quantity = int(upd.get('quantity', 1))
                    if not item_id:
                        continue
                    if quantity <= 0:
                        quantity = 1
                    try:
                        cart_item = CartItem.objects.get(id=item_id, cart=cart)
                        cart_item.quantity = quantity
                        cart_item.save()
                        updated_count += 1
                    except CartItem.DoesNotExist:
                        continue

                return JsonResponse({'success': True, 'updated_count': updated_count})

            elif action == 'bulk_remove':
                item_ids = data.get('item_ids', [])
                deleted_count = 0
                for item_id in item_ids:
                    try:
                        cart_item = CartItem.objects.get(id=item_id, cart=cart)
                        cart_item.delete()
                        deleted_count += 1
                    except CartItem.DoesNotExist:
                        continue

                return JsonResponse({'success': True, 'deleted_count': deleted_count})

            elif action == 'update_item':
                item_id = data.get('item_id')
                quantity = int(data.get('quantity', 1))
                if not item_id:
                    return JsonResponse({'success': False, 'message': 'Item ID is required'})
                if quantity <= 0:
                    quantity = 1
                try:
                    cart_item = CartItem.objects.get(id=item_id, cart=cart)
                    cart_item.quantity = quantity
                    cart_item.save()
                    return JsonResponse({'success': True, 'message': 'Cart item updated successfully'})
                except CartItem.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Cart item not found'})

            elif action == 'remove_item':
                item_id = data.get('item_id')
                if not item_id:
                    return JsonResponse({'success': False, 'message': 'Item ID is required'})
                try:
                    cart_item = CartItem.objects.get(id=item_id, cart=cart)
                    cart_item.delete()
                    return JsonResponse({'success': True, 'message': 'Cart item removed'})
                except CartItem.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Cart item not found'})

            elif action == 'update_cart_info':
                address_override = data.get('address_override')
                shipping_amount = data.get('shipping_amount')
                deposit_amount = data.get('deposit_amount')
                notes = data.get('notes')
                sales_person = data.get('sales_person')
                doc_ref = data.get('doc_ref')
                customer_code = data.get('customer_code')
                gross_weight = data.get('gross_weight')
                try:
                    cart.address_override = address_override
                    cart.shipping_amount = Decimal(str(shipping_amount or 0))
                    cart.deposit_amount = Decimal(str(deposit_amount or 0))
                    cart.notes = notes
                    cart.sales_person = sales_person
                    cart.doc_ref = doc_ref
                    cart.customer_code = customer_code
                    cart.gross_weight = Decimal(str(gross_weight or 0))
                    cart.save()
                    return JsonResponse({'success': True, 'message': 'Cart details updated'})
                except Exception as e:
                    return JsonResponse({'success': False, 'message': str(e)})
            else:
                return JsonResponse({'success': False, 'message': 'Invalid action'})

        return JsonResponse({'error': 'Method not allowed'}, status=405)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
@csrf_exempt
def customers_android_locked_count_api(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Only GET method allowed'}, status=405)
    count = Customer.objects.filter(locked=True).count()
    return JsonResponse({'locked_count': count})


@csrf_exempt
def customers_android_locked_ids_api(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Only GET method allowed'}, status=405)
    ids = list(Customer.objects.filter(locked=True).values_list('id', flat=True))
    return JsonResponse({'locked_ids': ids})
